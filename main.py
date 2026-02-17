import os
import re
import io
import csv
import time
import hashlib
import shutil
import requests
import openpyxl
from pathlib import Path
from typing import Optional
from fastapi import FastAPI, HTTPException, Header, Query, UploadFile, File, Body
from fastapi.responses import StreamingResponse, FileResponse, HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

app = FastAPI()

# --- CORS Configuration ---
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

print("\n" + "="*40)
print("XLSX-to-CSV Bridge Starting...")
print("="*40 + "\n")

@app.get("/")
async def root():
    return {"status": "ready", "message": "XLSX-to-CSV Bridge is running. Visit /test for the UI."}

@app.get("/test", response_class=HTMLResponse)
async def test_page():
    return """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>XLSX-to-CSV Bridge Test Bench</title>
    <style>
        :root {
            --primary: #2563eb;
            --bg: #f8fafc;
            --card: #ffffff;
            --text: #1e293b;
        }
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background-color: var(--bg);
            color: var(--text);
            display: flex;
            justify-content: center;
            align-items: center;
            min-height: 100vh;
            margin: 0;
        }
        .container {
            background: var(--card);
            padding: 2rem;
            border-radius: 1rem;
            box-shadow: 0 10px 25px -5px rgba(0, 0, 0, 0.1);
            width: 100%;
            max-width: 500px;
        }
        h1 {
            font-size: 1.5rem;
            margin-bottom: 1.5rem;
            text-align: center;
            color: var(--primary);
        }
        .field {
            margin-bottom: 1rem;
        }
        label {
            display: block;
            margin-bottom: 0.5rem;
            font-weight: 600;
        }
        input[type="text"], input[type="file"] {
            width: 100%;
            padding: 0.75rem;
            border: 1px solid #cbd5e1;
            border-radius: 0.5rem;
            box-sizing: border-box;
        }
        button {
            width: 100%;
            padding: 0.75rem;
            background-color: var(--primary);
            color: white;
            border: none;
            border-radius: 0.5rem;
            font-size: 1rem;
            font-weight: 600;
            cursor: pointer;
            transition: background 0.2s;
        }
        button:hover {
            background-color: #1d4ed8;
        }
        button:disabled {
            background-color: #94a3b8;
            cursor: not-allowed;
        }
        #status {
            margin-top: 1.5rem;
            padding: 1rem;
            border-radius: 0.5rem;
            display: none;
            white-space: pre-wrap;
            word-break: break-all;
        }
        .progress-group {
            margin-top: 1.5rem;
            display: none;
        }
        .progress-label {
            font-size: 0.875rem;
            margin-bottom: 0.5rem;
            display: flex;
            justify-content: space-between;
        }
        .progress-bar-container {
            width: 100%;
            height: 10px;
            background: #e2e8f0;
            border-radius: 5px;
            overflow: hidden;
            margin-bottom: 1rem;
        }
        .progress-bar-fill {
            height: 100%;
            background: var(--primary);
            width: 0%;
            transition: width 0.3s;
        }
        .progress-bar-fill.indeterminate {
            width: 100%;
            background: linear-gradient(90deg, #2563eb 25%, #60a5fa 50%, #2563eb 75%);
            background-size: 200% 100%;
            animation: move-bg 1.5s infinite linear;
        }
        @keyframes move-bg {
            0% { background-position: 200% 0; }
            100% { background-position: -200% 0; }
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>XLSX-to-CSV Bridge Tester</h1>
        
        <div class="field">
            <label for="bridgeUrl">Bridge Base URL</label>
            <input type="text" id="bridgeUrl" readonly value="">
        </div>

        <div class="field">
            <label for="apiKey">X-API-KEY</label>
            <input type="text" id="apiKey" value="a9F3kL8xQ2mZ7pR4tV6yH1nW5cX8bD0sE3uJ9gK2qL6zT4vY7">
        </div>

        <div class="field">
            <label for="fileInput">Select XLSX File (up to 100MB+)</label>
            <input type="file" id="fileInput" accept=".xlsx">
        </div>

        <div class="field">
            <label for="sheetName">Sheet Name (Optional)</label>
            <input type="text" id="sheetName" placeholder="Defaults to All Sheets Merged">
        </div>

        <button id="testBtn">Upload & Convert to CSV</button>

        <div class="progress-group" id="progressGroup">
            <div class="progress-label">
                <span>Uploading...</span>
                <span id="uploadPct">0%</span>
            </div>
            <div class="progress-bar-container">
                <div class="progress-bar-fill" id="uploadBar"></div>
            </div>

            <div class="progress-label" id="convertLabel" style="display:none">
                <span>Converting & Filtering (Please Wait)...</span>
            </div>
            <div class="progress-bar-container" id="convertBarContainer" style="display:none">
                <div class="progress-bar-fill indeterminate"></div>
            </div>
        </div>

        <div id="status"></div>
    </div>

    <script>
        // Auto-detect current host
        document.getElementById('bridgeUrl').value = window.location.origin;

        const testBtn = document.getElementById('testBtn');
        const statusDiv = document.getElementById('status');
        const progressGroup = document.getElementById('progressGroup');
        const uploadBar = document.getElementById('uploadBar');
        const uploadPct = document.getElementById('uploadPct');
        const convertLabel = document.getElementById('convertLabel');
        const convertBarContainer = document.getElementById('convertBarContainer');

        testBtn.addEventListener('click', async () => {
            const bridgeUrl = document.getElementById('bridgeUrl').value.trim();
            const apiKey = document.getElementById('apiKey').value.trim();
            const fileInput = document.getElementById('fileInput');
            const sheetName = document.getElementById('sheetName').value.trim();

            if (!fileInput.files.length) {
                showStatus('Please select a file first.', 'error');
                return;
            }

            const file = fileInput.files[0];
            const formData = new FormData();
            formData.append('file', file);

            const uploadUrl = `${bridgeUrl}/test-upload${sheetName ? `?sheet_name=${encodeURIComponent(sheetName)}` : ''}`;

            statusDiv.style.display = 'none';
            progressGroup.style.display = 'block';
            convertLabel.style.display = 'none';
            convertBarContainer.style.display = 'none';
            uploadBar.style.width = '0%';
            uploadPct.textContent = '0%';
            testBtn.disabled = true;

            const xhr = new XMLHttpRequest();
            xhr.open('POST', uploadUrl, true);
            xhr.setRequestHeader('X-API-KEY', apiKey);
            xhr.responseType = 'blob';

            xhr.upload.onprogress = (e) => {
                if (e.lengthComputable) {
                    const percent = Math.round((e.loaded / e.total) * 100);
                    uploadBar.style.width = percent + '%';
                    uploadPct.textContent = percent + '%';
                    
                    if (percent === 100) {
                        setTimeout(() => {
                            convertLabel.style.display = 'flex';
                            convertBarContainer.style.display = 'block';
                        }, 200);
                    }
                }
            };

            xhr.onload = async () => {
                progressGroup.style.display = 'none';
                if (xhr.status >= 200 && xhr.status < 300) {
                    const blob = xhr.response;
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    a.download = `${file.name.replace('.xlsx', '')}_filtered.csv`;
                    document.body.appendChild(a);
                    a.click();
                    a.remove();
                    showStatus('Success! Filtered CSV has been downloaded.', 'success');
                } else {
                    const reader = new FileReader();
                    reader.onload = () => showStatus(`Error ${xhr.status}: ${reader.result}`, 'error');
                    reader.readAsText(xhr.response);
                }
                testBtn.disabled = false;
            };

            xhr.onerror = () => {
                progressGroup.style.display = 'none';
                showStatus('Network Error during upload.', 'error');
                testBtn.disabled = false;
            };

            xhr.send(formData);
        });

        function showStatus(message, type) {
            statusDiv.textContent = message;
            statusDiv.className = type;
            statusDiv.style.display = 'block';
        }
    </script>
</body>
</html>
"""

# --- Configuration ---
API_KEY_NAME = "X-API-KEY"
REQUIRED_API_KEY = os.getenv("API_KEY")
CACHE_DIR = Path("/tmp/xlsx_cache")
CACHE_DIR.mkdir(exist_ok=True)

class ConversionRequest(BaseModel):
    drive_url: str
    sheet_name: str | None = None

def verify_api_key(x_api_key: str = Header(..., alias="X-API-KEY")):
    """Validates the X-API-KEY header."""
    if REQUIRED_API_KEY and x_api_key != REQUIRED_API_KEY:
        raise HTTPException(status_code=403, detail="Invalid API Key")
    return x_api_key

def extract_file_id(url: str) -> str:
    """Extracts the Google Drive file ID from various URL formats."""
    # Standard /d/ID format
    match = re.search(r"/d/([a-zA-Z0-9_-]+)", url)
    if match:
        return match.group(1)
    # uc?id=ID format
    match = re.search(r"id=([a-zA-Z0-9_-]+)", url)
    if match:
        return match.group(1)
    # open?id=ID format
    match = re.search(r"open\?id=([a-zA-Z0-9_-]+)", url)
    if match:
        return match.group(1)
    
    raise HTTPException(status_code=400, detail="Could not extract File ID from URL")

def parse_range_header(range_header: str, content_length: int):
    """Parses the Range header and returns start, end."""
    match = re.search(r"bytes=(\d+)-(\d*)", range_header)
    if match:
        start = int(match.group(1))
        end_str = match.group(2)
        end = int(end_str) if end_str else content_length - 1
        return start, end
    return 0, content_length - 1

def perform_conversion_from_url(file_id: str, sheet_name: str | None, cache_path: Path):
    """Downloads XLSX directly to disk and then converts to CSV on disk."""
    session = requests.Session()
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    }
    
    download_url = f"https://drive.google.com/uc?export=download&id={file_id}"
    tmp_xlsx = CACHE_DIR / f"{file_id}.xlsx"

    try:
        # Check if source XLSX is already cached and fresh (e.g. < 10 mins)
        if tmp_xlsx.exists() and (time.time() - tmp_xlsx.stat().st_mtime < 600):
            print(f"Using cached source XLSX for {file_id}")
        else:
            print(f"Downloading source XLSX for {file_id}")
            # 1. Stream download directly to disk (Critical for 100MB+ files)
            response = session.get(download_url, headers=headers, stream=True)
            if response.status_code == 200 and "confirm=" in response.text:
                confirm_match = re.search(r'confirm=([a-zA-Z0-9_-]+)', response.text)
                if confirm_match:
                    confirm_token = confirm_match.group(1)
                    download_url = f"https://drive.google.com/uc?export=download&id={file_id}&confirm={confirm_token}"
                    response = session.get(download_url, headers=headers, stream=True)
            
            response.raise_for_status()
            with open(tmp_xlsx, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)

        # 2. Convert from disk with filtering
        convert_xlsx_to_csv(tmp_xlsx, sheet_name, cache_path)
        
        # Note: We don't cleanup tmp_xlsx here anymore, so it's available for other sheet requests.
        # It will be overwritten/refreshed after 10 minutes.

    except Exception as e:
        if tmp_xlsx.exists(): tmp_xlsx.unlink()
        if isinstance(e, HTTPException): raise e
        raise HTTPException(status_code=500, detail=f"Conversion failed: {str(e)}")

def convert_xlsx_to_csv(xlsx_path: Path, sheet_name_req: str | None, csv_path: Path, target_val: str = "MRZ", date_val: str | None = None):
    """
    Memory-efficient conversion and FILTERING. 
    Finds "MRZ" rows in "Source_DC" or "DC" columns.
    Adds a Date column if date_val is provided.
    """
    try:
        wb = openpyxl.load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
        print(f"Workbook Details: Sheets={wb.sheetnames}")
        
        target_sheet = sheet_name_req if sheet_name_req and sheet_name_req.strip() else None

        sheets_to_process = []
        if target_sheet:
            target = target_sheet.strip().lower()
            found_ws = None
            if target_sheet in wb.sheetnames:
                found_ws = wb[target_sheet]
            else:
                for name in wb.sheetnames:
                    if name.strip().lower() == target:
                        found_ws = wb[name]
                        break
            if not found_ws:
                wb.close()
                raise HTTPException(status_code=400, detail=f"Sheet '{target_sheet}' not found")
            sheets_to_process.append(found_ws)
        else:
            for name in wb.sheetnames:
                sheets_to_process.append(wb[name])

        first_sheet = True
        with open(csv_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            
            for ws in sheets_to_process:
                print(f"Filtering sheet: {ws.title}")
                rows_iter = ws.iter_rows(values_only=True)
                
                try:
                    headers = next(rows_iter)
                except StopIteration:
                    continue

                if not headers: continue
                
                # Identify column indices
                col_map = {str(h).strip(): i for i, h in enumerate(headers) if h is not None}
                target_idx = col_map.get("Source_DC")
                if target_idx is None:
                    target_idx = col_map.get("DC")

                # Write headers only once
                if first_sheet:
                    header_list = list(headers)
                    if date_val: header_list.append("Date")
                    writer.writerow(header_list)
                    first_sheet = False

                if target_idx is not None:
                    for row in rows_iter:
                        if not row or len(row) <= target_idx: continue
                        
                        val = str(row[target_idx]).strip() if row[target_idx] is not None else ""
                        if val == target_val:
                            row_list = list(row)
                            if date_val: row_list.append(date_val)
                            writer.writerow(row_list)
        
        wb.close()
        print("Filtering & Conversion Complete.")
    except Exception as e:
        if isinstance(e, HTTPException): raise e
        raise HTTPException(status_code=500, detail=f"XLSX Parse Error: {str(e)}")

@app.post("/test-upload")
async def test_upload(
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Query(None),
    date_str: Optional[str] = Query(None),
    target_value: str = Query("MRZ"),
    x_api_key: Optional[str] = Header(None, alias="X-API-KEY")
):
    """Endpoint for manual testing via local file upload."""
    if REQUIRED_API_KEY and x_api_key != REQUIRED_API_KEY:
        raise HTTPException(status_code=403, detail="Unauthorized")

    file_id = f"test_{int(time.time())}"
    tmp_xlsx = CACHE_DIR / f"{file_id}.xlsx"
    cache_path = CACHE_DIR / f"{file_id}.csv"

    try:
        with open(tmp_xlsx, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        convert_xlsx_to_csv(tmp_xlsx, sheet_name, cache_path, target_val=target_value, date_val=date_str)
        
        if tmp_xlsx.exists(): tmp_xlsx.unlink()
        
        return FileResponse(
            path=cache_path,
            media_type="text/csv",
            filename="converted.csv"
        )
    except Exception as e:
        if tmp_xlsx.exists(): tmp_xlsx.unlink()
        raise HTTPException(status_code=500, detail=str(e))

async def handle_conversion_request(
    drive_url: Optional[str],
    sheet_name: Optional[str],
    api_key_provided: Optional[str],
    range_header: Optional[str],
    date_str: Optional[str] = None,
    target_value: str = "MRZ"
):
    # 1. Security Check
    if REQUIRED_API_KEY and api_key_provided != REQUIRED_API_KEY:
        print(f"Auth Failed: Provided={api_key_provided}")
        raise HTTPException(status_code=403, detail="Unauthorized")

    if not drive_url:
        raise HTTPException(status_code=400, detail="drive_url is required")

    file_id = extract_file_id(drive_url)
    
    # Clean sheet name for logging
    clean_sheet_log = sheet_name if sheet_name else "ALL_SHEETS_FILTERED"
    print(f"Request: File={file_id}, Sheet={clean_sheet_log}, Target={target_value}")

    # Create a unique cache key based on params
    cache_key_raw = f"{file_id}_{sheet_name}_{target_value}_{date_str}"
    cache_key = hashlib.md5(cache_key_raw.encode()).hexdigest()
    cache_path = CACHE_DIR / f"{cache_key}.csv"

    # 3. Check Cache
    if not cache_path.exists() or (time.time() - cache_path.stat().st_mtime > 600):
        print(f"Starting Filter & Conversion for {file_id}")
        perform_conversion_from_url_with_filter(file_id, sheet_name, cache_path, target_value, date_str)

    total_size = cache_path.stat().st_size
    print(f"Response Size: {total_size} bytes")

    # 4. Handle Range Request or Full Response
    if range_header:
        start, end = parse_range_header(range_header, total_size)
        if end >= total_size: end = total_size - 1
        if start > end: raise HTTPException(status_code=416, detail="Range Not Satisfiable")

        chunk_size = end - start + 1
        def iterfile():
            with open(cache_path, "rb") as f:
                f.seek(start)
                yield f.read(chunk_size)

        return StreamingResponse(
            iterfile(),
            status_code=206,
            media_type="text/csv",
            headers={
                "Content-Range": f"bytes {start}-{end}/{total_size}",
                "Content-Length": str(chunk_size),
                "Accept-Ranges": "bytes",
                "Content-Disposition": f"attachment; filename={file_id}.csv"
            }
        )
    else:
        return FileResponse(
            path=cache_path,
            media_type="text/csv",
            filename=f"{file_id}.csv",
            headers={"Accept-Ranges": "bytes"}
        )

@app.get("/convert")
async def convert_get(
    drive_url: Optional[str] = Query(None),
    sheet_name: Optional[str] = Query(None),
    api_key: Optional[str] = Query(None),
    date_str: Optional[str] = Query(None),
    target_value: str = Query("MRZ"),
    x_api_key: Optional[str] = Header(None, alias="X-API-KEY"),
    range_header: Optional[str] = Header(None, alias="Range")
):
    return await handle_conversion_request(drive_url, sheet_name, x_api_key or api_key, range_header, date_str, target_value)

@app.post("/convert")
async def convert_post(
    request_data: Optional[ConversionRequest] = Body(None),
    drive_url: Optional[str] = Query(None),
    sheet_name: Optional[str] = Query(None),
    api_key: Optional[str] = Query(None),
    date_str: Optional[str] = Query(None),
    target_value: str = Query("MRZ"),
    x_api_key: Optional[str] = Header(None, alias="X-API-KEY"),
    range_header: Optional[str] = Header(None, alias="Range")
):
    url = drive_url or (request_data.drive_url if request_data else None)
    sn = sheet_name or (request_data.sheet_name if request_data else None)
    return await handle_conversion_request(url, sn, x_api_key or api_key, range_header, date_str, target_value)

def perform_conversion_from_url_with_filter(file_id: str, sheet_name: str | None, cache_path: Path, target_val: str, date_val: str | None):
    """Refactored version of perform_conversion_from_url to carry filter params."""
    session = requests.Session()
    headers = {"User-Agent": "Mozilla/5.0"}
    download_url = f"https://drive.google.com/uc?export=download&id={file_id}"
    tmp_xlsx = CACHE_DIR / f"{file_id}.xlsx"

    try:
        if not tmp_xlsx.exists() or (time.time() - tmp_xlsx.stat().st_mtime > 600):
            print(f"Downloading source XLSX for {file_id}")
            response = session.get(download_url, headers=headers, stream=True)
            # Handle confirmation page
            if "confirm=" in response.text:
                token = re.search(r'confirm=([a-zA-Z0-9_-]+)', response.text).group(1)
                download_url += f"&confirm={token}"
                response = session.get(download_url, headers=headers, stream=True)
            response.raise_for_status()
            with open(tmp_xlsx, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)

        convert_xlsx_to_csv(tmp_xlsx, sheet_name, cache_path, target_val, date_val)
    except Exception as e:
        if tmp_xlsx.exists(): tmp_xlsx.unlink()
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
